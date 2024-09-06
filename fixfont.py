import os
import openpyxl
import unicodedata
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QThread, Signal


class ExcelProcessorThread(QThread):
    # 定义信号，用于传递结果和状态
    send_result = Signal(str)

    # 定义一个 CJK 部首字符到标准汉字的映射
    mapping_dict = {
        '⻩': '黄', '⻊': '足', '⻋': '车', '⻔': '门', '女': '女',
        '⼀': '一', '⼁': '丨', '⼂': '丶', '⼃': '丿', '⼄': '乙',
        '⼅': '亅', '⼆': '二', '⼇': '亠', '⼈': '人', '⼉': '儿',
        '⼊': '入', '⼋': '八', '⼌': '冂', '⼍': '冖', '⼎': '冫',
        '⼏': '几', '⼐': '凵', '⼑': '刀', '⼒': '力', '⼓': '勹',
        '⼔': '匕', '⼕': '匚', '⼖': '匸', '⼗': '十', '⼘': '卜',
        '⼙': '卩', '⼚': '厂', '⼛': '厶', '⼜': '又', '⼝': '口',
        '⼞': '囗', '⼟': '土', '⼠': '士', '⼡': '夂', '⼢': '夊',
        '⼣': '夕', '⼤': '大', '⼥': '女', '⼦': '子', '⼧': '宀',
        '⼨': '寸', '⼩': '小', '⼪': '尢', '⼫': '尸', '⼬': '屮',
        '⼭': '山', '⼮': '巛', '⼯': '工', '⼰': '己', '⼱': '巾',
        '⼲': '干', '⼳': '幺', '⼴': '广', '⼵': '廴', '⼶': '廾',
        '⼷': '弋', '⼸': '弓', '⼹': '彐', '⼺': '彡', '⼻': '彳',
        '⼼': '心', '⼽': '戈', '⼾': '戶', '⼿': '手', '⽀': '支',
        '⽁': '攴', '⽂': '文', '⽃': '斗', '⽄': '斤', '⽅': '方',
        '⽆': '无', '⽇': '日', '⽈': '曰', '⽉': '月', '⽊': '木',
        '⽋': '欠', '⽌': '止', '⽍': '歹', '⽎': '殳', '⽏': '毋',
        '⽐': '比', '⽑': '毛', '⽒': '氏', '⽓': '气', '⽔': '水',
        '⽕': '火', '⽖': '爪', '⽗': '父', '⽘': '爻', '⽙': '爿',
        '⽚': '片', '⽛': '牙', '⽜': '牛', '⽝': '犬', '⽞': '玄',
        '⽟': '玉', '⽠': '瓜', '⽡': '瓦', '⽢': '甘', '⽣': '生',
        '⽤': '用', '⽥': '田', '⽦': '疋', '⽧': '疒', '⽨': '癶',
        '⽩': '白', '⽪': '皮', '⽫': '皿', '⽬': '目', '⽭': '矛',
        '⽮': '矢', '⽯': '石', '⽰': '示', '⽱': '禸', '⽲': '禾',
        '⽳': '穴', '⽴': '立', '⽵': '竹', '⽶': '米', '⽷': '糸',
        '⽸': '缶', '⽹': '网', '⽺': '羊', '⽻': '羽', '⽼': '老',
        '⽽': '而', '⽾': '耒', '⽿': '耳', '⾀': '聿', '⾁': '肉',
        '⾂': '臣', '⾃': '自', '⾄': '至', '⾅': '臼', '⾆': '舌',
        '⾇': '舛', '⾈': '舟', '⾉': '艮', '⾊': '色', '⾋': '艸',
        '⾌': '虍', '⾍': '虫', '⾎': '血', '⾏': '行', '⾐': '衣',
        '⾑': '襾', '⾒': '见', '⾓': '角', '⾔': '言', '⾕': '谷',
        '⾖': '豆', '⾗': '豕', '⾘': '豸', '⾙': '贝', '⾚': '赤',
        '⾛': '走', '⾜': '足', '⾝': '身', '⾞': '车', '⾟': '辛',
        '⾠': '辰', '⾡': '辵', '⾢': '邑', '⾣': '酉', '⾤': '釆',
        '⾥': '里', '⾦': '金', '⾧': '长', '⾨': '门', '⾩': '阜',
        '⾪': '隶', '⾫': '隹', '⾬': '雨', '⾭': '青', '⾮': '非',
        '⾯': '面', '⾰': '革', '⾱': '韋', '⾲': '韭', '⾳': '音',
        '⾴': '頁', '⾵': '风', '⾶': '飞', '⾷': '食', '⾸': '首',
        '⾹': '香', '⾺': '马', '⾻': '骨', '⾼': '高', '⾽': '髟',
        '⾾': '鬥', '⾿': '鬯', '⿀': '鬲', '⿁': '鬼', '⿂': '鱼',
        '⿃': '鸟', '⿄': '卤', '⿅': '鹿', '⿆': '麦', '⿇': '麻',
        '⿈': '黄', '⿉': '黍', '⿊': '黑', '⿋': '黹', '⿌': '黽',
        '⿍': '鼎', '⿎': '鼓', '⿏': '鼠', '⿐': '鼻', '⿑': '齐',
        '氵': '水', '木': '木', '火': '火', '土': '土', '金': '金',
        '㐩': '乐', '㐪': '亚', '㐫': '亨', '㐬': '亭', '㐭': '亮',
        '㐮': '亲', '㐯': '义', '㐰': '举', '㐱': '传', '㐲': '伤',
        '㐳': '伦', '㐴': '似', '㐵': '作', '㐶': '你', '㐷': '佘',
        '㐸': '佛', '㐹': '位', '㐺': '体', '㐻': '作', '㐼': '保',
        '㐽': '便', '㐾': '信', '㐿': '倒', '㑀': '修', '㑁': '假',
        '㑂': '伟', '㑃': '健', '㑄': '偏', '㑅': '做', '㑆': '停',
        '㑇': '倩', '㑈': '像', '㑉': '僧', '㑊': '僧', '㑋': '僵',
        '㑌': '免', '㑍': '党', '㑎': '入', '㑏': '农', '㑐': '公',
        '㑑': '共', '㑒': '兵', '㑓': '内', '㑔': '全', '㑕': '八',
        '㑖': '六', '㑗': '兴', '㑘': '冲', '㑙': '冯', '㑚': '冻',
        '㑛': '减', '㑜': '出', '㑝': '函', '㑞': '刚', '㑟': '刀',
        '㑠': '分', '㑡': '初', '㑢': '别', '㑣': '利', '㑤': '制',
        '㑥': '刻', '㑦': '剪', '㑧': '力', '㑨': '功', '㑩': '劳',
        '㑪': '动', '㑫': '办', '㑬': '务', '㑭': '化', '㑮': '医',
        '㑯': '南', '㑰': '博', '㑱': '卡', '㑲': '参', '㑳': '单',
        '㑴': '双', '㑵': '台', '㑶': '右', '㑷': '叶', '㑸': '召',
        '㑹': '合', '㑺': '吉', '㑻': '名', '㑼': '后', '㑽': '君',
        '㑾': '和', '㑿': '味', '㒀': '品', '㒁': '唐', '㒂': '回',
        '㒃': '因', '㒄': '国', '㒅': '园', '㒆': '图', '㒇': '圆',
        '㒈': '圣', '㒉': '士', '㒊': '寿', '㒋': '帝', '㒌': '帅',
        '色': '色', '⻘': '青', '尺': '尺', '寸': '寸',
        '米': '米', '分': '分', '斤': '斤', '两': '两'
        # 你可以在这里添加更多的部首字符和对应的汉字
    }

    def __init__(self, parent=None):
        super().__init__(parent)
        self.file_path = None
        self.output_path = None

    def run(self):
        try:
            # 使用openpyxl读取Excel文件，保留格式
            workbook = load_workbook(self.file_path)

            # 遍历工作表
            for sheet_name in workbook.sheetnames:
                # 使用pandas读取当前工作表数据
                df = pd.read_excel(self.file_path, sheet_name=sheet_name)

                # 保留列的顺序
                original_columns = df.columns

                # 使用 DataFrame 的 apply 和 map 来对所有单元格进行处理
                df_modified = df.apply(
                    lambda col: col.map(lambda cell: self.clean_and_replace(cell, self.mapping_dict)))

                # 获取对应的工作表
                worksheet = workbook[sheet_name]

                # 遍历合并单元格并记录哪些单元格是合并单元格的左上角
                merged_cells = []
                for merged_range in worksheet.merged_cells.ranges:
                    merged_cells.append(merged_range.bounds)

                # 遍历修改后的数据框，将值写回到openpyxl的worksheet中
                for row_idx, row in df_modified.iterrows():
                    for col_idx, value in enumerate(row):
                        col_letter = get_column_letter(col_idx + 1)
                        cell = worksheet[f'{col_letter}{row_idx + 2}']

                        # 检查当前单元格是否是合并区域中的非左上角单元格
                        in_merged_range = False
                        for (min_col, min_row, max_col, max_row) in merged_cells:
                            if min_col <= col_idx + 1 <= max_col and min_row <= row_idx + 2 <= max_row:
                                if col_idx + 1 != min_col or row_idx + 2 != min_row:
                                    in_merged_range = True
                                    break

                        if not in_merged_range:
                            # 只有在不属于合并区域的非左上角单元格时，才能写入
                            cell.value = value

            # 在保存之前，保持隐藏行和列的状态
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                for row_dimension in worksheet.row_dimensions.values():
                    if row_dimension.hidden:  # 检查行是否隐藏
                        worksheet.row_dimensions[row_dimension.index].hidden = True
                for col_dimension in worksheet.column_dimensions.values():
                    if col_dimension.hidden:  # 检查列是否隐藏
                        worksheet.column_dimensions[col_dimension.index].hidden = True

            # 保存修改后的工作簿
            workbook.save(self.output_path)
            print(f"成功将文件保存为: {self.output_path}")
            self.send_result.emit("Excel已修复成功")

        except FileNotFoundError:
            print(f"错误: 找不到文件 {self.file_path}")
            self.send_result.emit(f"Excel修复失败: 找不到文件 {self.file_path}")
        except Exception as e:
            print(f"发生错误: {e}")
            self.send_result.emit(f"Excel修复失败: {e}")

    # 新增的函数，用于去除空格和替换字符
    def clean_and_replace(self, cell_value, mapping_dict):
        """
        先去除单元格内容中的空格，再替换CJK Unicode字符或部首字符为标准汉字。

        参数:
        cell_value (str): 单元格的内容。
        mapping_dict (dict): CJK 部首字符到标准汉字的映射字典。

        返回:
        str: 清理并替换后的内容。
        """
        if isinstance(cell_value, str):
            # 去除字符串中的所有空格
            cleaned_value = cell_value.replace(" ", "").strip()
            # 调用原有的替换方法
            return self.replace_unicode_cjk_characters(cleaned_value, mapping_dict)
        return cell_value

    # 判断是否为CJK部首字符
    def is_cjk_radical(self, char):
        return char in self.mapping_dict

    # 判断是否为Unicode字符（CJK等其他范围的字符）
    def is_unicode_character(self, char):
        try:
            char_name = unicodedata.name(char)
            return "CJK" in char_name or "UNICODE" in char_name
        except ValueError:
            return False

    def replace_unicode_cjk_characters(self, cell_value, mapping_dict):
        if isinstance(cell_value, str):
            new_value = []
            for char in cell_value:
                if self.is_cjk_radical(char):
                    new_value.append(mapping_dict[char])
                elif self.is_unicode_character(char):
                    new_value.append(char)
                else:
                    new_value.append(char)
            return ''.join(new_value)
        return cell_value

    def set_path(self, file_path, output_path):
        self.file_path = file_path
        self.output_path = output_path


# 主函数用于启动线程
def main():
    root_path = r"D:\ChromeDownload"
    input_excel_name = "壹品公馆9-6-2"
    input_excel_path = f'{root_path}/{input_excel_name}.xlsx'  # 输入文件路径
    output_excel_path = f'{root_path}/{input_excel_name}_repaired.xlsx'  # 输出文件路径

    # 检查输入文件是否存在
    if not os.path.exists(input_excel_path):
        print(f"输入文件不存在: {input_excel_path}")
        return

    # 创建线程并启动
    thread = ExcelProcessorThread(input_excel_path, output_excel_path)

    # 连接信号到槽函数
    thread.finished.connect(lambda msg: print(msg))
    thread.error.connect(lambda msg: print(msg))

    # 启动线程
    thread.start()


if __name__ == "__main__":
    main()
